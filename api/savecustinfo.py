from openpyxl import load_workbook

filename = 'D:\\Python_work\\walnuts_api\\api_test\\resources\\test.xlsx'
vip_img01 = 'D:\\Python_work\\walnuts_api\\api_test\\resources\\vip_img01.txt'
vip_img02 = 'D:\\Python_work\\walnuts_api\\api_test\\resources\\vip_img02.txt'
vip_img03 = 'D:\\Python_work\\walnuts_api\\api_test\\resources\\vip_img03.txt'

def writeExcel(sheetname,custinfo):
    '''写入excel'''
    wb = load_workbook(filename)
    ws1 = wb.create_sheet(sheetname)  #默认最后一个
    ws1.append(custinfo)
    wb.save(filename)


def readExcel(sheetname):
    '''读取excel'''
    custinfo = []
    wb = load_workbook(filename)
    sheet = wb[sheetname]
    rows = sheet.max_row
    cols = sheet.max_column
    for c in range(1,cols+1):
        custinfo.append(sheet.cell(1, c).value)
    return custinfo


def save_vipimg_log(filename,repones):
    with open(filename, 'w', encoding='utf-8') as f_obj:
        f_obj.write(repones)


def read_log_img(filename):
    with open(filename) as file_object:
        contents = file_object.read()
        return contents


if __name__ == "__main__":
    vip_img01 = 'D:\\Python_work\\walnuts_api\\api_test\\resources\\vip_img01.txt'
    vip_img02 = 'D:\\Python_work\\walnuts_api\\api_test\\resources\\vip_img02.txt'
    vip_img03 = 'D:\\Python_work\\walnuts_api\\api_test\\resources\\vip_img03.txt'

    print(app[0])

